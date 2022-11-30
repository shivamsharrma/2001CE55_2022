import pandas as pd
import math
from datetime import datetime
import glob
from io import BytesIO 
import streamlit as st
start_time = datetime.now()
st.title("Gui version of tut 7")
uploaded_file= st.file_uploader("Choose a file")
#Help
mod=int(st.number_input('Enter the mod value:'))
if uploaded_file is not None and mod!=0:
	def proj_octant_gui():


		
		dp=pd.read_excel(uploaded_file)
	
		dp=dp.fillna(0) 
		octant_name_id_mapping = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}
		avg_u=dp['U'].mean() #finding avg of u using mean function
		avg_v=dp['V'].mean() #finding avg of v using mean function
		avg_w=dp['W'].mean() #finding avg of w using mean function

		dp.loc[0,"U Avg"]=avg_u  #assigning avg_u 0th row
		dp.loc[0,"V Avg"]=avg_v  #assigning avg_v 0th row
		dp.loc[0,"W Avg"]=avg_w  #assigning avg_w 0th row

		dp["U'=U - U avg"]=dp['U']-dp.loc[0,"U Avg"]  #finding U'
		dp["V'=V - V avg"]=dp['V']-dp.loc[0,'V Avg']  #finding V'
		dp["W'=W - W avg"]=dp['W']-dp.loc[0,'W Avg']  #finding W'
		octa=[]
		count1=0   #to count all co-ordinate in octant 1
		count2=0   #to count all co-ordinate in octant 2
		count3=0   #to count all co-ordinate in octant 3
		count4=0   #to count all co-ordinate in octant 4
		count5=0   #to count all co-ordinate in octant -1
		count6=0   #to count all co-ordinate in octant -2
		count7=0   #to count all co-ordinate in octant -3
		count8=0   #to count all co-ordinate in octant -4


		for row in range(len(dp)):  #for loop used to append octat value to variable octa
			border_a=dp.loc[row,"U'=U - U avg"]      #assigning row th index of U column to a
			b=dp.loc[row,"V'=V - V avg"]      #assigning row th index of V column to b
			c=dp.loc[row,"W'=W - W avg"]      #assigning row th index of W column to c
			
			if border_a >= 0 and b >= 0 and c >= 0:
				octa.append(1) 
				count1=count1+1
			elif border_a < 0 and b >= 0 and c >= 0:
				octa.append(2)
				count2=count2+1
				
			elif border_a < 0 and b < 0 and c >= 0:
				octa.append(3)
				count3=count3+1
				
			elif border_a >= 0 and b < 0 and c >= 0:
				octa.append(4)
				count4=count4+1
			elif border_a >= 0 and b >= 0 and c < 0:
				octa.append(-1)
				count5=count5+1
			elif border_a < 0 and b >= 0 and c < 0:
				octa.append(-2)
				count6=count6+1
			elif border_a < 0 and b < 0 and c < 0:
				octa.append(-3)
				count7=count7+1
			elif border_a >= 0 and b < 0 and c < 0:
				octa.append(-4)
				count8=count8+1
		dp['Octant']=octa #assiging all octa value to a column octant
		dp.loc[2," "]="User input" #made a column without heading and at 1st index assigning vaue to it
		dp["Octant ID"]=' ' #made a column with heading octant ID
		dp.loc[1,'Octant ID']='Overall Count'  #at 0th index of octant ID provided other value

		dp.loc[2,'Octant ID']="Mod {}".format(mod)  #used format specifier which assign value to mod on basis of input


		dp.loc[1,'1']=count1  # making column of heading 1 and assigning it count at 0th index
		dp.loc[1,'-1']=count5 #making column of heading -1 and assigning it count at 0th index
		dp.loc[1,'2']=count2  #making column of heading 2 and assigning it count at 0th index
		dp.loc[1,'-2']=count6 #making column of heading -2 and assigning it count at 0th index
		dp.loc[1,'3']=count3  #making column of heading 3 and assigning it count at 0th index
		dp.loc[1,'-3']=count7 #making column of heading -3 and assigning it count at 0th index
		dp.loc[1,'4']=count4  # making column of heading 4 and assigning it count at 0th index
		dp.loc[1,'-4']=count8 #making column of heading -4 and assigning it count at 0th index
		list1=[count1,count5,count2,count6,count3,count7,count4,count8]
		list1.sort(reverse=True)
		z=math.ceil(len(dp)/mod) #taking a ceiling function so that we can take upper value for any decimal value(3.44->4) to upper integer
		k=1
		dp.loc[0," 1"]="Rank 1"# at 0th index of new column adding a rank1
		dp.loc[0," -1"]="Rank 2"#at 0th index of new column adding a rank2
		dp.loc[0," 2"]="Rank 3"#at 0th index of new column adding a rank3
		dp.loc[0," -2"]="Rank 4"#at 0th index of new column adding a rank4
		dp.loc[0," 3"]="Rank 5"#at 0th index of new column adding a rank5
		dp.loc[0," -3"]="Rank 6"#at 0th index of new column adding a rank6
		dp.loc[0," 4"]="Rank 7"#at 0th index of new column adding a rank7
		dp.loc[0," -4"]="Rank 8"#at 0th index of new column adding a rank8
		dp.loc[0,"        "]="Rank1 OctantID"#at 0th index of new column adding a 
		dp.loc[0,"         "]="Rank1 Octant Name"#at 0th index of new column adding a 
		count11=0# made this count to know about no. of times 1 in comin raank 1
		count12=0  # made this count to know about no. of times -1 in comin raank 1
		count21=0# made this count to know about no. of times 2 in comin raank 1
		count22=0# made this count to know about no. of times -2 in comin raank 1
		count31=0# made this count to know about no. of times 3in comin raank 1
		count32=0# made this count to know about no. of times -3 in comin raank 1
		count41=0# made this count to know about no. of times 4 in comin raank 1
		count42=0# made this count to know about no. of times -4 in comin raank 1
		if(st.button('Compute')):
			while k<z:   #iterating over all octant value and finding in different mod no. of different octant present
				first=k*mod   #used to seefirst value of mod range
				last=((k+1)*mod)-1  #used to find last value of mod range
				
				if last>len(dp): #for last case when last value is great  3er than aur range
					last=len(dp)-1
					
				dp.loc[k+2,"Octant ID"]="{}-{}".format(first,last) 
				p=first
				q1=q2=q3=q4=q5=q6=q7=q8=0   # to count again all octant vale present in given mod range

				while p<=last:     #iterating 
					if dp.loc[p,"Octant"]==1:
						q1=q1+1
					elif dp.loc[p,"Octant"]==2:
						q2=q2+1
					elif dp.loc[p,"Octant"]==3:
						q3=q3+1
					elif dp.loc[p,"Octant"]==4:
						q4=q4+1
					elif dp.loc[p,"Octant"]==-1:
						q5=q5+1  
					elif dp.loc[p,"Octant"]==-2:
						q6=q6+1
					elif dp.loc[p,"Octant"]==-3:
						q7=q7+1
					elif dp.loc[p,"Octant"]==-4:
						q8=q8+1
					p=p+1
				list2=[q1,q5,q2,q6,q3,q7,q4,q8]  # made a list of all count and sorting it then we will know which number should come at rank 1 
				list2.sort(reverse=True)
				i=0
				while i in range(8):# in this while loop we are checking for which no. is rank one ,two  so on . and we are also telling in new column which index is rank 1 in that row and assainging it name
					if(list2[i]==q1):
						dp.loc[k+2," 1"]=i+1
						if(i==0):
							dp.loc[k+2,"        "]="1"
							dp.loc[k+2,"         "]="Internal outward interaction"
							count11+=1
					elif(list2[i]==q5):
						dp.loc[k+2," -1"]=i+1
						if(i==0):
							dp.loc[k+2,"        "]="-1"
							dp.loc[k+2,"         "]="External outward interaction"
							count12+=1
					elif(list2[i]==q2):
						dp.loc[k+2," 2"]=i+1
						if(i==0):
							dp.loc[k+2,"        "]="2"
							dp.loc[k+2,"         "]="External Ejection"
							count21+=1
					elif(list2[i]==q6):
						dp.loc[k+2," -2"]=i+1
						if(i==0):
							dp.loc[k+2,"        "]="-2"
							dp.loc[k+2,"         "]="Internal Ejection"
							count22+=1
					elif(list2[i]==q3):
						dp.loc[k+2," 3"]=i+1
						if(i==0):
							dp.loc[k+2,"        "]="3"
							dp.loc[k+2,"         "]="External inward interaction"
							count31+=1
					elif(list2[i]==q7):
						dp.loc[k+2," -3"]=i+1
						if(i==0):
							dp.loc[k+2,"        "]="-3"
							dp.loc[k+2,"         "]="Internal inward interaction"
							count32+=1
					elif(list2[i]==q4):
						dp.loc[k+2," 4"]=i+1
						if(i==0):
							dp.loc[k+2,"        "]="4"
							dp.loc[k+2,"         "]="Internal sweep"
							count41+=1
							
					elif(list2[i]==q8):
						dp.loc[k+2," -4"]=i+1
						if(i==0):
							dp.loc[k+2,"        "]="-4"
							dp.loc[k+2,"         "]="External sweep"
							count42+=1
					i+=1
				dp.loc[k+2,'1']=q1  #assigning count of no. of 1 in given range of mod
				dp.loc[k+2,'2']=q2  #assigning count of no. of 2 in given range of mod
				dp.loc[k+2,'3']=q3  #assigning count of no. of 3 in given range of mod
				dp.loc[k+2,'4']=q4  #assigning count of no. of 4 in given range of mod 
				dp.loc[k+2,'-1']=q5  #assigning count of no. of -1 in given range of mod
				dp.loc[k+2,'-2']=q6  #assigning count of no. of -2 in given range of mod
				dp.loc[k+2,'-3']=q7  #assigning count of no. of -3 in given range of mod
				dp.loc[k+2,'-4']=q8   #assigning count of no. of -4 in given range of mod       
				k=k+1

			k+=4

			i=0
			while i in range(8): # doing same stuff as above but this is only for total count 
				if(list1[i]==count1):
					dp.loc[1," 1"]=i+1
					if(i==0):
						dp.loc[1,"        "]="1"
						dp.loc[1,"         "]="Internal outward interaction"
				elif(list1[i]==count5):
					dp.loc[1," -1"]=i+1
					if(i==0):
						dp.loc[1,"        "]="-1"
						dp.loc[1,"         "]="External outward interaction"
				elif(list1[i]==count2):
					dp.loc[1," 2"]=i+1
					if(i==0):
						dp.loc[1,"        "]="2"
						dp.loc[1,"         "]="External Ejection"
				elif(list1[i]==count6):
					dp.loc[1," -2"]=i+1
					if(i==0):
						dp.loc[1,"        "]="-2"
						dp.loc[1,"         "]="Internal Ejection"
				elif(list1[i]==count3):
					dp.loc[1," 3"]=i+1
					if(i==0):
						dp.loc[1,"        "]="3"
						dp.loc[1,"         "]="External inward interaction"
				elif(list1[i]==count7):
					dp.loc[1," -3"]=i+1
					if(i==0):
						dp.loc[1,"        "]="-3"
						dp.loc[1,"         "]="Internal inward interaction"
				elif(list1[i]==count4):
					dp.loc[1," 4"]=i+1
					if(i==0):
						dp.loc[1,"        "]="4"
						dp.loc[1,"         "]="Internal sweep"
				elif(list1[i]==count8):
					dp.loc[1," -4"]=i+1
					if(i==0):
						dp.loc[1,"        "]="-4"
						dp.loc[1,"         "]="External sweep"
				i+=1

			dp.loc[k," 4"]="OCTANT ID"        #MAKING A header at kth position from top as octant id
			dp.loc[k," -4"]="Octant Name"
			dp.loc[k,"        "]="Count of Rank 1 Mod Values"
			k+=1
			z=0
			list3=[count11,count12,count21,count22,count31,count32,count41,count42] ##making a list which is storing the count of no of time each element came as rank ``
			for i,j in octant_name_id_mapping.items(): # here we are telling what is mean by 1,-1,2,-2,3,-3,4,-4and making new column to tell how much times it come rank 1
				dp.loc[k," 4"]=i
				dp.loc[k," -4"]=j
				dp.loc[k,"        "]=list3[z]
				k+=1
				z+=1
			
			dp['we']=' '
			k=0
			dp.loc[k,'Octant ID ']="Overall Transition Count"   #printing head
			dp.loc[k+3,'we']="From"                              
			dp.loc[k+1,"1 "]="To"
			dp.loc[k+2,'Octant ID ']="Count" 
			dp.loc[k+2,'1 ']='1 '    #printing head of column to know different transition count
			dp.loc[k+2,'-1 ']='-1 '  #printing head of column to know different transition count
			dp.loc[k+2,'2 ']="2 "      #printing head of column to know different transition count
			dp.loc[k+2,'-2 ']="-2 "    #printing head of column to know different transition count
			dp.loc[k+2,'3 ']="3 "      #printing head of column to know different transition count
			dp.loc[k+2,'-3 ']="-3 "    #printing head of column to know different transition count
			dp.loc[k+2,'4 ']="4 "      #printing head of column to know different transition count
			dp.loc[k+2,'-4 ']="-4 "   #printing head of column to know different transition count

			k=k+3
			dp.loc[k,'Octant ID ']='1 '     ##printing head of row to know different transition count
			dp.loc[k+1,'Octant ID ']='-1 '  ##printing head of row to know different transition count
			dp.loc[k+2,'Octant ID ']='2 '  ##printing head of row to know different transition count
			dp.loc[k+3,'Octant ID ']='-2 '  ##printing head of row to know different transition count
			dp.loc[k+4,'Octant ID ']='3 '  ##printing head of row to know different transition count
			dp.loc[k+5,'Octant ID ']='-3 '  ##printing head of row to know different transition count
			dp.loc[k+6,'Octant ID ']='4 '  ##printing head of row to know different transition count
			dp.loc[k+7,'Octant ID ']='-4 '  ##printing head of row to know different transition count
			############################################
			a=["1 ","-1 ",'2 ','-2 ' ,'3 ','-3 ','4 ','-4 ']  #making a list so that we can define the 8*8 row column as 0 to count the transition
			try:
				for i in a:           # for loop to access the element of list a
					for j in range(8):  # making variable j so that we can define 8*8 row column as 0
						dp.loc[j+k,i]=0  # since we want 8*8 row column 0  from row k from top therefore added k
			except:
				print("error may be due to out of range or due to wrong indexing ")
			l=0
			try:
				while(l+1<len(dp)):            #this while loop is for counting total  transition count 
					i=dp.loc[l,'Octant']        #taking the lth row element of octant column
					j=str(dp.loc[l+1,'Octant'])  #taking the l+1th row element of octant column ans defining it in string because we defined column1,-1,2,-2,3,-3,4,-4 as string
					if(j=='1'):
							x='1 '
					elif j=='-1':
						x='-1 '
					elif j=='2':
						x='2 '
					elif j=='-2':
						x='-2 '
					elif j=='3':
						x='3 '
					elif j=='-3':
						x='-3 '
					elif j=='4':
						x='4 '
					elif j=='-4':
						x='-4 '
					if i==1:                      #now seeing if lth element is 1 then going at row k (since row k represent 1)and at column l+1th  and increasing its count
						dp.loc[k,x]=dp.loc[k,x]+1
					elif i==-1:
						dp.loc[k+1,x]=dp.loc[k+1,x]+1#now seeing if lth element is -1 then going at row k (since row k represent -1)and at column l+1th  and increasing its count
					elif i==2:
						dp.loc[k+2,x]=dp.loc[k+2,x]+1#now seeing if lth element is 2 then going at row k (since row k represent 1)and at column l+1th  and increasing its count
					elif i==-2:
						dp.loc[k+3,x]=dp.loc[k+3,x]+1 #now seeing if lth element is -2 then going at row k (since row k represent 1)and at column l+1th  and increasing its count
					elif i==3:
						dp.loc[k+4,x]=dp.loc[k+4,x]+1 #now seeing if lth element is 3 then going at row k (since row k represent 1)and at column l+1th  and increasing its count
					elif i==-3:
						dp.loc[k+5,x]=dp.loc[k+5,x]+1  #now seeing if lth element is -3 then going at row k (since row k represent 1)and at column l+1th  and increasing its count
					elif i==4:
						dp.loc[k+6,x]=dp.loc[k+6,x]+1  #now seeing if lth element is 4 then going at row k (since row k represent 1)and at column l+1th  and increasing its count
					elif i==-4:
						dp.loc[k+7,x]=dp.loc[k+7,x]+1    #now seeing if lth element is -4 then going at row k (since row k represent 1)and at column l+1th  and increasing its count
					l=l+1
			except:
				print('error may be due to string to int conversion')
					

			k=k+11     #now increassing the row no. so that we can also count mod transition count
			##################################################
			r=0
			try:
				z=math.ceil(len(dp)/mod)   #using cealing funtion to take greater value of any decimal no. so that we can know iin how many part we need to count
			except:
				print("error may be due to mod =0")
			try:
				while(r<z):             #using while loop which runs same time as no. of part we need to divide
					first=r*mod         #this line is written so that we can take range-> intial  value
					last=((r+1)*mod)-1    #this line is written so that we can take range ->last value  (here not taking last value-1)because we need the transition count between last and 2nd lastpoint
					if(last>len(dp)-1):   # this line is writtern so that we can avoid the error of range value asn aur range is less tha 30000 so loop  last ccan nt be greater than len(dp)
						last=len(dp)-1
					dp.loc[k,"Octant ID "]="Mod Transition Count"
					dp.loc[k+3,"we"]="From"
					dp.loc[k+1,"Octant ID "]="{}-{}".format(first,last)   # this is used to write range (first,last)
					dp.loc[k+1,"1 "]='To'
					dp.loc[k+2,'Octant ID ']="Count"
					dp.loc[k+2,'1 ']='1 '  
					dp.loc[k+2,'-1 ']='-1 ' 
					dp.loc[k+2,'2 ']="2 " 
					dp.loc[k+2,'-2 ']="-2 " 
					dp.loc[k+2,'3 ']="3 " 
					dp.loc[k+2,'-3 ']="-3 " 
					dp.loc[k+2,'4 ']="4 "  
					dp.loc[k+2,'-4 ']="-4 "
					k+=3
					dp.loc[k,'Octant ID ']='1 '     ##printing head of row to know different transition count
					dp.loc[k+1,'Octant ID ']='-1 '  ##printing head of row to know different transition count
					dp.loc[k+2,'Octant ID ']='2 '  ##printing head of row to know different transition count
					dp.loc[k+3,'Octant ID ']='-2 '  ##printing head of row to know different transition count
					dp.loc[k+4,'Octant ID ']='3 '  ##printing head of row to know different transition count
					dp.loc[k+5,'Octant ID ']='-3 '  ##printing head of row to know different transition count
					dp.loc[k+6,'Octant ID ']='4 '  ##printing head of row to know different transition count
					dp.loc[k+7,'Octant ID ']='-4 '  ##printing head of row to know different transition count
						
					for i in a:   
						for j in range(8):  #this loop is again writtern so that we can make 8*8row column as 0
							dp.loc[j+k,i]=0
					x=''
					while first<last+1 if last!=len(dp)-1 else first<last:      #now from range (first to last ) finding the no. of transition count taking place
						i=dp.loc[first,'Octant']    # taking first row  element of octant colummn
						j=str(dp.loc[first+1,'Octant'])  #taking first+1 th row element of octant column
						if(j=='1'):
							x='1 '
						elif j=='-1':
							x='-1 '
						elif j=='2':
							x='2 '
						elif j=='-2':
							x='-2 '
						elif j=='3':
							x='3 '
						elif j=='-3':
							x='-3 '
						elif j=='4':
							x='4 '
						elif j=='-4':
							x='-4 '
						if i==1:
							dp.loc[k,x]=dp.loc[k,x]+1
						elif i==-1:
							dp.loc[k+1,x]=dp.loc[k+1,x]+1
						elif i==2:
							dp.loc[k+2,x]=dp.loc[k+2,x]+1
						elif i==-2:
							dp.loc[k+3,x]=dp.loc[k+3,x]+1
						elif i==3:
							dp.loc[k+4,x]=dp.loc[k+4,x]+1
						elif i==-3:
							dp.loc[k+5,x]=dp.loc[k+5,x]+1
						elif i==4:
							dp.loc[k+6,x]=dp.loc[k+6,x]+1
						elif i==-4:
							dp.loc[k+7,x]=dp.loc[k+7,x]+1    
						first=first+1
					
					r=r+1
					k=k+11
			except:
				print("error may be due to out of range ,int to string converter")
			dp[" e0"]=" "
			dp.loc[0,"e1"]="Count"  #made a column and assiging the 0th row as count
			dp.loc[0,"e2"]="Longest Subsequence Length"   #made a column and assigned th 0th row as longest common subsequence
			dp.loc[0,"e3"]="Count"   # made 3rd extra column and again assigned its 0th element as count
			dp.loc[1,"e1"]="1"     #in new 1st column and different row  assinged string as 1,-1,2,-2,3,-3,-4,4
			dp.loc[2,"e1"]="-1"
			dp.loc[3,"e1"]="2"   #'3 space e1   1 space e2 2 space e3   5 space e5 6 space e6  7 space e7
			dp.loc[4,"e1"]="-2"
			dp.loc[5,"e1"]="3"
			dp.loc[6,"e1"]="-3"
			dp.loc[7,"e1"]="4"
			dp.loc[8,"e1"]="-4"
			dp['e4']=' '
			dp.loc[0,"e5"]="Count"
			dp.loc[0,"e6"]="Longest Subsequence Length"
			dp.loc[0,"e7"]="Count"
			f=0
			g=0
			e=1

			a=[1,-1,2,-2,3,-3,4,-4] #made a list  so that we can traverse element in all octant element
			for  g in range(8): #made a for loop to iterate over all 
				count1=0
				max_count1=0
				ans=1
				t=0
				f=0
				while f<len(dp):  #this while loop is to iterate over octant value and print table given in tut3 
					
					if dp.loc[f,'Octant']==a[g]:   # checking if  fth value of octant column is matching with our list gth element and then counting its occurance
						count1+=1     
					else:
						if max_count1==count1 and count1!=0:       # if this count is  equal to max count then increasing the count of ans 
							ans+=1
						elif count1>max_count1:              # if count is greater than max count then we writting then max count as present count and  making ans as 1
							max_count1=count1
							ans=1
						count1=0       # now making count as 0 as we need to iterate again
					f+=1
				dp.loc[g+1,"e2"]=max_count1   #assigning the max count at g+1th row of c column
				dp.loc[1+g,"e3"]=ans             #assigning the ans at g+1th row of c column
				dp.loc[e,"e5"]=a[g]            # from here making a table for tut 4 containing the octant value by using list
				dp.loc[e,"e6"]=max_count1       # here giving the  max xcount in front of octant value 
				dp.loc[e,"e7"]=ans
				e+=1
				dp.loc[e,"e5"]="Time"        # making another row and assingning value as time , from ,to
				dp.loc[e,"e6"]="From"
				dp.loc[e,"e7"]="To"
				e+=1
				f=0
				while f<len(dp):    # this while loop is to  see the interval in which we will be getting the longest subsequence of that octant value ans also help in printing them in table made 
					if dp.loc[f,'Octant']==a[g]:   # again checking for max count 
						if(count1==0):
							t=f              # noting down the iterator value form where our longest subsequence is suppossed to start
							
						count1+=1
					else:
						if max_count1==count1 and count1!=0:
							ans+=1
							dp.loc[e,"e6"]=dp.loc[t,"T"]   # printing down the intial time at which aur subse quence started only if it get end 
							dp.loc[e,"e7"]=dp.loc[f-1,"T"] # printing down the final time at which our subsequence is  ended
							e+=1
						elif count1>max_count1:
							max_count1=count1
							ans=1
						count1=0
					f+=1
				g+=1
			dp.to_excel("output.xlsx",index=False)
			from datetime import datetime
			start_time = datetime.now()
			import openpyxl
			wb = openpyxl.load_workbook('output.xlsx')
			sheet = wb.active
			from openpyxl.styles import PatternFill,Border,Side
			for row_1 in sheet.iter_rows(min_row=2, max_row=7, min_col=22, max_col=29):
				yellow = "00FFFF00"
				for cell_1 in row_1:
					if cell_1.value==1:
						cell_1.fill = PatternFill(start_color=yellow, end_color=yellow,fill_type = "solid")
			print(sheet['AD3'].value==1)
			border_a=Side(border_style='thin',color="00000000")
			row_count=9
			column_count=4
			row_position=9
			column_position=27
			border=Border(top=border_a,bottom=border_a,left=border_a,right=border_a)
			for i in range (row_position, row_position+row_count):
				for j in range (column_position,column_count+column_position) :
					sheet.cell(row=i+1, column=j+1).border=border
					if i==row_position:
						sheet. cell (row=i+1,column=j+1).border=border
					if i==row_position+row_count-1:
						sheet.cell(row=i+1,column=j+1).border=border
			row_count=7
			column_count=19
			row_position=0
			column_position=12
			border=Border(top=border_a,bottom=border_a,left=border_a,right=border_a)
			for i in range (row_position, row_position+row_count):
				for j in range (column_position,column_count+column_position) :
					sheet.cell(row=i+1, column=j+1).border=border
					if i==row_position:
						sheet. cell (row=i+1,column=j+1).border=border
					if i==row_position+row_count-1:
						sheet.cell(row=i+1,column=j+1).border=border
			
			row_count=25
			column_count=3
			row_position=1
			column_position=46
			border=Border(top=border_a,bottom=border_a,left=border_a,right=border_a)
			for i in range (row_position, row_position+row_count):
				for j in range (column_position,column_count+column_position) :
					sheet.cell(row=i+1, column=j+1).border=border
					if i==row_position:
						sheet. cell (row=i+1,column=j+1).border=border
					if i==row_position+row_count-1:
						sheet.cell(row=i+1,column=j+1).border=border

			row_count=9
			column_count=3
			row_position=1
			column_position=42
			border=Border(top=border_a,bottom=border_a,left=border_a,right=border_a)
			for i in range (row_position, row_position+row_count):
				for j in range (column_position,column_count+column_position) :
					sheet.cell(row=i+1, column=j+1).border=border
					if i==row_position:
						sheet. cell (row=i+1,column=j+1).border=border
					if i==row_position+row_count-1:
						sheet.cell(row=i+1,column=j+1).border=border
			
			

			row_count=9
			column_count=9
			row_position=3
			column_position=32
			dis=5
			#vary
			Color_count=4+1
			for k in range(Color_count):
				for i in range (row_position, row_position+row_count):
					for j in range (column_position,column_count+column_position) :
						sheet.cell(row=i+1, column=j+1).border=border
						if i==row_position:
							sheet. cell (row=i+1,column=j+1).border=border
						if i==row_position+row_count-1:
							sheet.cell(row=i+1,column=j+1).border=border
				row_position+=row_count+dis

			row_count=8
			column_count=8
			row_position=4
			column_position=34
			dis=6
			#vary
			Color_count=4+1

			list_1=[[],[],[],[],[],[],[],[]]
			for k in range(Color_count):
				for i in range (row_position, row_position+row_count):
					list_1=[[],[],[],[],[],[],[],[]]
					x=0
					for j in range (column_position,column_count+column_position) :
						list_1[x].append(sheet.cell(row=i,column=j).value)
						list_1[x].append(int(i))
						list_1[x].append(int(j))
						x+=1
						border_a=max(list_1)
					sheet.cell(row=border_a[1],column=border_a[2]).fill=PatternFill(start_color=yellow, end_color=yellow,fill_type = "solid")
				row_position+=row_count+dis

			wb.save('output.xlsx')
			st.dataframe(dp.head(10))
			with open('output.xlsx',"rb") as f:
				file=f.read()
			
			st.download_button(label='🔻Download excel',data=file,file_name="output.xlsx")
    
	end_time = datetime.now()
	print('Duration of Program Execution: {}'.format(end_time - start_time))

	proj_octant_gui()
##Read all the excel files in a batch format from the input/ folder. Only xlsx to be allowed
##Save all the excel files in a the output/ folder. Only xlsx to be allowed
## output filename = input_filename[_octant_analysis_mod_5000].xlsx , ie, append _octant_analysis_mod_5000 to the original filename. 

###Code

from platform import python_version
ver = python_version()

# if ver == "3.8.10":
# print("Correct Version Installed")
# else:
# print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")



#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))