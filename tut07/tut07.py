
import csv
from datetime import datetime

start_time = datetime.now()
import math  # importing math
from sqlite3 import Time
import openpyxl
import pandas as pd  # importing pandas


def octant_analysis(mod=5000):
   #code to read input csv file and store that in dataframe
    
    df=pd.read_excel('input\\1.0.xlsx') 

    #takingh mean of each 'U','V','W' column
    df1=df["U"].mean()
    df2=df["V"].mean()
    df3=df["W"].mean()

    #printing average values
    df.loc[0,"U Avg"]=df1
    df.loc[0,"V Avg"]=df2
    df.loc[0,"W Avg"]=df3

    
    df["U'=U- U Avg"]=df["U"]-df1
    df["V'=V- V Avg"]=df["V"]-df2
    df["W'=W- W Avg"]=df["W"]-df3

    oct=[] #creating empty array
    #initializing count of every octant is zero
    count1=0
    count2=0
    count3=0
    count4=0
    count5=0
    count6=0
    count7=0
    count8=0

    
    #initializing loop 
    try:
        for row in range(len(df)):
            x=df.loc[row,"U'=U- U Avg"] #calculating difference between 'U' and 'U Avg'
            y=df.loc[row,"V'=V- V Avg"] #calculating difference between 'V' and 'V Avg'
            z=df.loc[row,"W'=W- W Avg"] #calculating difference between 'W' and 'W Avg'
            #checking in which octant the value lies and appending according value to oct array
            if x >= 0 and y >= 0 and z >= 0:
                oct.append(1)
                count1+=1
                
            elif x < 0 and y >= 0 and z >= 0:
                oct.append(2)
                count2+=1
                
            elif x < 0 and y < 0 and z >= 0:
                oct.append(3)
                count3+=1
                
            elif x >= 0 and y < 0 and z >= 0:
                oct.append(4)
                count4+=1
                
            elif x >= 0 and y >= 0 and z < 0:
                oct.append(-1)
                count5+=1
                
            elif x < 0 and y >= 0 and z < 0:
                oct.append(-2)
                count6+=1
                
            elif x < 0 and y < 0 and z < 0:
                oct.append(-3)
                count7+=1

            elif x >= 0 and y < 0 and z < 0:
                oct.append(-4)
                count8+=1
    except:
        print("There is error in counting the number of octant")

    df["Octant"]=oct #copying oct values to octant column
    df[" "]="" #intiliazing a empty column
    df.loc[1," "]="User Input" #asigning string to 1st row of empty column
    df["Octant ID"]='' #creating octant id column
    df.loc[0,"Octant ID"]="Overall Count" #printing "Overall Count" to 0th row of "Octant ID" column
    df["1"]=''
    df.loc[0,"1"]=count1 #printing overall count of 1
    df["-1"]=''
    df.loc[0,"-1"]=count5 #printing overall count of -1
    df["2"]=''
    df.loc[0,"2"]=count2 #printing overall count of 2
    df["-2"]=''
    df.loc[0,"-2"]=count6 #printing overall count of -2
    df["3"]=''
    df.loc[0,"3"]=count3 #printing overall count of 3
    df["-3"]=''
    df.loc[0,"-3"]=count7 #printing overall count of -3
    df["4"]=''
    df.loc[0,"4"]=count4 #printing overall count of 4
    df["-4"]=''
    df.loc[0,"-4"]=count8 #printing overall count of -4
    
    
    array=[count1,count2,count3,count4,count5,count6,count7,count8,]
    array.sort(reverse=True)
    df[1]=df[-1]=df[2]=df[-2]=df[3]=df[-3]=df[4]=df[-4]=" "
    df["Rank1 Octant ID"]=""
    df["Rank1 Octant Name"]=""
    octant_name={1:"Internal outward interaction",-1:"External outward interaction",2:"External Ejection",
    -2:"Internal Ejection",3:"External inward interaction",-3:"Internal inward interaction",
    4:"Internal sweep",-4:"External sweep"}
    
    rank=0
    try:
        for i in array:
            if(i==df.loc[0,"1"]):
                df.loc[0,1]=rank+1
                if(rank==0):
                    df.loc[0,"Rank1 Octant ID"]=1
                    df.loc[0,"Rank1 Octant Name"]=octant_name[1]
            elif(i==df.loc[0,"-1"]):
                df.loc[0,-1]=rank+1
                if(rank==0):
                    df.loc[0,"Rank1 Octant ID"]=-1
                    df.loc[0,"Rank1 Octant Name"]=octant_name[-1]
            elif(i==df.loc[0,"2"]):
                df.loc[0,2]=rank+1
                if(rank==0):
                    df.loc[0,"Rank1 Octant ID"]=2
                    df.loc[0,"Rank1 Octant Name"]=octant_name[2]
            elif(i==df.loc[0,"-2"]):
                df.loc[0,-2]=rank+1
                if(rank==0):
                    df.loc[0,"Rank1 Octant ID"]=-2
                    df.loc[0,"Rank1 Octant Name"]=octant_name[-2]
            elif(i==df.loc[0,"3"]):
                df.loc[0,3]=rank+1
                if(rank==0):
                    df.loc[0,"Rank1 Octant ID"]=3
                    df.loc[0,"Rank1 Octant Name"]=octant_name[3]
            elif(i==df.loc[0,"-3"]):
                df.loc[0,-3]=rank+1
                if(rank==0):
                    df.loc[0,"Rank1 Octant ID"]=-3
                    df.loc[0,"Rank1 Octant Name"]=octant_name[-3]
            elif(i==df.loc[0,"4"]):
                df.loc[0,4]=rank+1
                if(rank==0):
                    df.loc[0,"Rank1 Octant ID"]=4
                    df.loc[0,"Rank1 Octant Name"]=octant_name[4]
            elif(i==df.loc[0,"-4"]):
                df.loc[0,-4]=rank+1
                if(rank==0):
                    df.loc[0,"Rank1 Octant ID"]=-4
                    df.loc[0,"Rank1 Octant Name"]=octant_name[-4]
            rank+=1
    except:
        print("There is error in printing the rank of octant") 


    df.loc[1,"Octant ID"]="mod {}".format(mod) #printing mod value in 1st row of "Octant ID" column
    n=math.ceil(len(df)/mod) #calculating number of intervals
    #here cor represents count of rank
    cor1=cor2=cor3=cor4=cor5=cor6=cor7=cor8=0
    i=0
    try:
        while i<n:
            a=i*mod #lower value of interval
            b=((i+1)*mod)-1 #maximum value of interval
            if b>len(df):
                b=len(df)-1
            df.loc[i+2,"Octant ID"]="{}-{}".format(a,b) #printing intervals
            j=a 
            
            #initializing count of each octant to zero in the particular interval
            c1=0
            c2=0
            c3=0
            c4=0
            c5=0
            c6=0
            c7=0
            c8=0
            while j<=b:
                
                #counting the number of octants in the particular interval
                if df.loc[j,"Octant"]==1:
                    c1+=1
                elif df.loc[j,"Octant"]==2:
                    c2+=1
                elif df.loc[j,"Octant"]==3:
                    c3+=1
                elif df.loc[j,"Octant"]==4:
                    c4+=1
                elif df.loc[j,"Octant"]==-1:
                    c5+=1
                elif df.loc[j,"Octant"]==-2:
                    c6+=1
                elif df.loc[j,"Octant"]==-3:
                    c7+=1
                elif df.loc[j,"Octant"]==-4:
                    c8+=1
                j+=1
            
            #printing the count of each octants
            df.loc[i+2,"1"]=c1
            df.loc[i+2,"-1"]=c5
            df.loc[i+2,"2"]=c2
            df.loc[i+2,"-2"]=c6
            df.loc[i+2,"3"]=c3
            df.loc[i+2,"-3"]=c7
            df.loc[i+2,"4"]=c4
            df.loc[i+2,"-4"]=c8
            rank=0
            array=[c1,c2,c3,c4,c5,c6,c7,c8]
            array.sort(reverse=True)
            for element in array:
                if(element==df.loc[i+2,"1"]):
                    df.loc[i+2,1]=rank+1
                    if(rank==0):
                        df.loc[i+2,"Rank1 Octant ID"]=1
                        df.loc[i+2,"Rank1 Octant Name"]=octant_name[1]
                        cor1+=1
                elif(element==df.loc[i+2,"-1"]):
                    df.loc[i+2,-1]=rank+1
                    if(rank==0):
                        df.loc[i+2,"Rank1 Octant ID"]=-1
                        df.loc[i+2,"Rank1 Octant Name"]=octant_name[-1]
                        cor2+=1
                elif(element==df.loc[i+2,"2"]):
                    df.loc[i+2,2]=rank+1
                    if(rank==0):
                        df.loc[i+2,"Rank1 Octant ID"]=2
                        df.loc[i+2,"Rank1 Octant Name"]=octant_name[2]
                        cor3+=1
                elif(element==df.loc[i+2,"-2"]):
                    df.loc[i+2,-2]=rank+1
                    if(rank==0):
                        df.loc[i+2,"Rank1 Octant ID"]=-2
                        df.loc[i+2,"Rank1 Octant Name"]=octant_name[-2]
                        cor4+=1
                elif(element==df.loc[i+2,"3"]):
                    df.loc[i+2,3]=rank+1
                    if(rank==0):
                        df.loc[i+2,"Rank1 Octant ID"]=3
                        df.loc[i+2,"Rank1 Octant Name"]=octant_name[3]
                        cor5+=1
                elif(element==df.loc[i+2,"-3"]):
                    df.loc[i+2,-3]=rank+1
                    if(rank==0):
                        df.loc[i+2,"Rank1 Octant ID"]=-3
                        df.loc[i+2,"Rank1 Octant Name"]=octant_name[-3]
                        cor6+=1
                elif(element==df.loc[i+2,"4"]):
                    df.loc[i+2,4]=rank+1
                    if(rank==0):
                        df.loc[i+2,"Rank1 Octant ID"]=4
                        df.loc[i+2,"Rank1 Octant Name"]=octant_name[4]
                        cor7+=1
                elif(element==df.loc[i+2,"-4"]):
                    df.loc[i+2,-4]=rank+1
                    if(rank==0):
                        df.loc[i+2,"Rank1 Octant ID"]=-4
                        df.loc[i+2,"Rank1 Octant Name"]=octant_name[-4]
                        cor8+=1
                rank+=1

            i+=1
    except:
        print("There is error in printing the rank for mod values")
    i+=3
    df.loc[i,4]="Octant ID"
    df.loc[i,-4]="Octant Name" 
    df.loc[i,"Rank1 Octant ID"]="Count of Rank 1 Mod Values"
    i+=1

    rank_list={1:cor1,-1:cor2,2:cor3,-2:cor4,3:cor5,-3:cor6,4:cor7,-4:cor8}
    list=[1,-1,2,-2,3,-3,4,-4]

    for j in list:
        df.loc[i,4]=j
        df.loc[i,-4]=octant_name[j]
        df.loc[i,"Rank1 Octant ID"]=rank_list[j]
        i+=1

    df.rename(columns={1:'Rank of 1',-1:'Rank of -1',2:'Rank of 2',-2:'Rank of -2',3:'Rank of 3',
    -3:'Rank of -3',4:'Rank of 4',-4:'Rank of 4'},inplace=True)
    df['    ']=''
    df['  ']=''
    
    i=0
    df['Octant#']=''
    df[' 1']=df[' -1']=df[' 2']=df[' -2']=df[' 3']=df[' -3']=df[' 4']=df[' -4']='  '
    df.loc[i,'Octant#']=1
    df.loc[i+1,'Octant#']=-1
    df.loc[i+2,'Octant#']=2
    df.loc[i+3,'Octant#']=-2
    df.loc[i+4,'Octant#']=3
    df.loc[i+5,'Octant#']=-3
    df.loc[i+6,'Octant#']=4
    df.loc[i+7,'Octant#']=-4
    k=1 #initializing the row from 1
    list1= [' 1',' -1',' 2',' -2',' 3',' -3',' 4',' -4']
    for l in list1:
        for j in range (8):
            df.loc[j+i,l]=0 #assigning 0 to each value of matrix
    while k<len(df): #loop for counting the transition 
            prev=df.loc[k-1,"Octant"] #storing the previous octant value
            curr=' '+str(df.loc[k,"Octant"]) ##storing the current octant value
            #checking the transition values and increament the corresponding transition
            if prev==1:
                df.loc[i,curr]+=1
            elif prev==-1:
                df.loc[i+1,curr]+=1
            elif prev==2:
                df.loc[i+2,curr]+=1
            elif prev==-2:
                df.loc[i+3,curr]+=1
            elif prev==3:
                df.loc[i+4,curr]+=1
            elif prev==-3:
                df.loc[i+5,curr]+=1
            elif prev==4:
                df.loc[i+6,curr]+=1
            elif prev==-4:
                df.loc[i+7,curr]+=1
            k+=1    
    
    i=i+11

    c=0 #initiliazing the no. of interval from zero 
    while c<n:
        a=c*mod #lower value of interval
        b=((c+1)*mod) #upper bound
        if b>len(df):
            b=len(df)
        df.loc[i,'Octant#']="Mod Transition Count"
        i+=1
        df.loc[i,'Octant#']="{}-{}".format(a,b-1) #printing intervals
        df.loc[i,' 1']="To"
        i+=1
        df.loc[i,'Octant#']="Octant#"
        df.loc[i+1,'  ']="From"
        #Assigning octants in a column of matrix
        df.loc[i,' 1']=1
        df.loc[i,' -1']=-1
        df.loc[i,' 2']=2
        df.loc[i,' -2']=-2
        df.loc[i,' 3']=3
        df.loc[i,' -3']=-3
        df.loc[i,' 4']=4
        df.loc[i,' -4']=-4
        i+=1
        #Assigning octants in a row of matrix
        df.loc[i,'Octant#']=1
        df.loc[i+1,'Octant#']=-1
        df.loc[i+2,'Octant#']=2
        df.loc[i+3,'Octant#']=-2
        df.loc[i+4,'Octant#']=3
        df.loc[i+5,'Octant#']=-3
        df.loc[i+6,'Octant#']=4
        df.loc[i+7,'Octant#']=-4
        
        
        list1= [' 1',' -1',' 2',' -2',' 3',' -3',' 4',' -4']
        for l in list1:
            for j in range (8):
                df.loc[j+i,str(l)]=0
        j=a
        if b==len(df):
            b=b-1
        while j<b: #loop for calculating transition for each interval (a-b)
            prev=df.loc[j,"Octant"] #previous value of octant 
            curr=' '+str(df.loc[j+1,"Octant"]) #current value of octant

            #now checking the condition how the transition goes
            if prev==1:
                df.loc[i,curr]+=1
            elif prev==-1:
                df.loc[i+1,curr]+=1
            elif prev==2:
                df.loc[i+2,curr]+=1
            elif prev==-2:
                df.loc[i+3,curr]+=1
            elif prev==3:
                df.loc[i+4,curr]+=1
            elif prev==-3:
                df.loc[i+5,curr]+=1
            elif prev==4:
                df.loc[i+6,curr]+=1
            elif prev==-4:
                df.loc[i+7,curr]+=1
            j+=1    
        i=i+11
         
        c+=1
        df['   ']='    '
         #intiliazing a empty column\
        list=[1,-1,2,-2,3,-3,4,-4]
        df["Octant##"]="" #creating the column count
        j=0
        while j<8:
            df.loc[j,"Octant##"]=list[j] #printing all the values of list in count column
            j+=1
        df["Longest subsequence length"]="" #creating the longest subsequence length column which stores count of longest subsequence
        df["count"]=" " #creating count array which stores count of longest subsequence
        count1=count2=count3=count4=count5=count6=count7=count8=0 #initializing longest subsequence length to zero
        c1=c2=c3=c4=c5=c6=c7=c8=0 #secondary variable which count the longest subsequence count
        # #l_count represents the count of longest subsequence length
        l_count1=l_count2=l_count3=l_count4=l_count5=l_count6=l_count7=l_count8=0 
        for y in range (len(df)-1):
            curr_oct=df.loc[y,"Octant"] #storing current octant value
            next_oct=df.loc[y+1,"Octant"] #storing next octant value
            if curr_oct==next_oct:
                if curr_oct==1:
                    c1+=1
                elif curr_oct==-1:
                    c2+=1
                elif curr_oct==2:
                    c3+=1
                elif curr_oct==-2:
                    c4+=1
                elif curr_oct==3:
                    c5+=1
                elif curr_oct==-3:
                    c6+=1
                elif curr_oct==4:
                    c7+=1
                elif curr_oct==-4:
                    c8+=1
            else:
                if curr_oct==1:
                    c1+=1
                    if c1==count1:
                        l_count1+=1
                    elif c1>count1:
                        l_count1=1
                    count1=max(count1,c1)
                    c1=0
                elif curr_oct==-1:
                    c2+=1
                    if c2==count2:
                        l_count2+=1
                    elif c2>count2:
                        l_count2=1
                    count2=max(count2,c2)
                    c2=0
                elif curr_oct==2:
                    c3+=1
                    if c3==count3:
                        l_count3+=1
                    elif c3>count3:
                        l_count3=1
                    count3=max(count3,c3)
                    c3=0
                elif curr_oct==-2:
                    c4+=1
                    if c4==count4:
                        l_count4+=1
                    elif c4>count4:
                        l_count4=1
                    count4=max(count4,c4)
                    c4=0
                elif curr_oct==3:
                    c5+=1
                    if c5==count5:
                        l_count5+=1
                    elif c5>count5:
                        l_count5=1
                    count5=max(count5,c5)
                    c5=0
                elif curr_oct==-3:
                    c6+=1
                    if c6==count6:
                        l_count6+=1
                    elif c6>count6:
                        l_count6=1
                    count6=max(count6,c6)
                    c6=0
                elif curr_oct==4:
                    c7+=1
                    if c7==count7:
                        l_count7+=1
                    elif c7>count7:
                        l_count7=1
                    count7=max(count7,c7)
                    c7=0
                elif curr_oct==-4:
                    c8+=1
                    if c8==count8:
                        l_count8+=1
                    elif c8>count8:
                        l_count8=1
                    count8=max(count8,c8)
                    c8=0
        
    df.loc[0,"Longest subsequence length"]=count1
    df.loc[1,"Longest subsequence length"]=count2
    df.loc[2,"Longest subsequence length"]=count3
    df.loc[3,"Longest subsequence length"]=count4
    df.loc[4,"Longest subsequence length"]=count5
    df.loc[5,"Longest subsequence length"]=count6
    df.loc[6,"Longest subsequence length"]=count7
    df.loc[7,"Longest subsequence length"]=count8  


    # locating count of longest subsequence length in the data in its corresponding row
    df.loc[0,"count"]= l_count1 
    df.loc[1,"count"]= l_count2
    df.loc[2,"count"]= l_count3
    df.loc[3,"count"]= l_count4
    df.loc[4,"count"]= l_count5
    df.loc[5,"count"]= l_count6
    df.loc[6,"count"]= l_count7
    df.loc[7,"count"]= l_count8
    df['     ']='  '
    df["Octant### "]='  '  #creating another empty column
    df["Longest subsequence length "]='  ' #creating another empty column
    df["Count "]='  ' #creating another empty column 

    index=0 #variable which hover on every octant
    row =0

    for i in list: 
            count=0 #initializing count of every octant from zero
            df.loc[row,"Octant### "]=i #locating octant in new octant column
            df.loc[row,"Longest subsequence length "]=df.loc[index,"Longest subsequence length"] #printing Longest subsequence length in new Longest subsequence length column
            df.loc[row,"Count "]=df.loc[index,"count"]
            row+=1
            df.loc[row,"Octant### "]='Time'
            df.loc[row,"Longest subsequence length "]='From'
            df.loc[row,"Count "]='To'
            row+=1
            #loop which checks for each octant and its count
            for j in range (len(df)):
                if(df.loc[j,"Octant"]==i):
                    count+=1
                else:
                    if(count==df.loc[index,"Longest subsequence length"]):
                        #printing start time for octant
                        df.loc[row,"Longest subsequence length "]=df.loc[j-df.loc[index,"Longest subsequence length"],"T"]
                        #printing end time for octant
                        df.loc[row,"Count "]=df.loc[j-1,"T"]
                        row+=1
                        count=0
                    else:
                        count=0
            index+=1
    df.to_excel('output.xlsx',index=False)

    wb = openpyxl.load_workbook('output.xlsx')
    sheet = wb.active
    from openpyxl.styles import PatternFill,Border,Side
    for rows in sheet.iter_rows(min_row=2, max_row=7, min_col=22, max_col=29):
        yellow = "00FFFF00"
        for cell in rows:
            if cell.value==1:
                cell.fill = PatternFill(start_color=yellow, end_color=yellow,fill_type = "solid")
    a=Side(border_style='thin',color="00000000")
    rownum=6
    colnum=19
    rowloc=1
    colloc=12
    border=Border(top=a,bottom=a,left=a,right=a)
    for i in range (rowloc, rowloc+rownum):
        for j in range (colloc,colnum+colloc) :
            sheet.cell(row=i+1, column=j+1).border=border
            if i==rowloc:
                sheet. cell (row=i+1,column=j+1).border=border
            if i==rowloc+rownum-1:
                sheet.cell(row=i+1,column=j+1).border=border
    rownum=9
    colnum=3
    rowloc=8
    colloc=27
    border=Border(top=a,bottom=a,left=a,right=a)
    for i in range (rowloc, rowloc+rownum):
        for j in range (colloc,colnum+colloc) :
            sheet.cell(row=i+1, column=j+1).border=border
            if i==rowloc:
                sheet. cell (row=i+1,column=j+1).border=border
            if i==rowloc+rownum-1:
                sheet.cell(row=i+1,column=j+1).border=border
    rownum=9
    colnum=3
    rowloc=0
    colloc=43
    border=Border(top=a,bottom=a,left=a,right=a)
    for i in range (rowloc, rowloc+rownum):
        for j in range (colloc,colnum+colloc) :
            sheet.cell(row=i+1, column=j+1).border=border
            if i==rowloc:
                sheet. cell (row=i+1,column=j+1).border=border
            if i==rowloc+rownum-1:
                sheet.cell(row=i+1,column=j+1).border=border
    # rownum=25+vary
    rownum=25
    colnum=3
    rowloc=0
    colloc=47
    border=Border(top=a,bottom=a,left=a,right=a)
    for i in range (rowloc, rowloc+rownum):
        for j in range (colloc,colnum+colloc) :
            sheet.cell(row=i+1, column=j+1).border=border
            if i==rowloc:
                sheet. cell (row=i+1,column=j+1).border=border
            if i==rowloc+rownum-1:
                sheet.cell(row=i+1,column=j+1).border=border
    rownum=9
    colnum=9
    rowloc=0
    colloc=33
    dis=5
    #vary
    T=4+1
    for k in range(T):
        for i in range (rowloc, rowloc+rownum):
            for j in range (colloc,colnum+colloc) :
                sheet.cell(row=i+1, column=j+1).border=border
                if i==rowloc:
                    sheet. cell (row=i+1,column=j+1).border=border
                if i==rowloc+rownum-1:
                    sheet.cell(row=i+1,column=j+1).border=border
        rowloc+=rownum+dis
    rownum=8
    colnum=8
    rowloc=2
    colloc=35
    dis=6
    #vary
    T=4+1
    l=[[],[],[],[],[],[],[],[]]
    for k in range(T):
        for i in range (rowloc, rowloc+rownum):
            l=[[],[],[],[],[],[],[],[]]
            x=0
            for j in range (colloc,colnum+colloc) :
                l[x].append(sheet.cell(row=i,column=j).value)
                l[x].append(i)
                l[x].append(j)
                x+=1
                a=max(l)
                print(l)
            sheet.cell(row=a[1],column=a[2]).fill=PatternFill(start_color=yellow, end_color=yellow,fill_type = "solid")
        rowloc+=rownum+dis
    wb.save("color.xlsx")
    
mod=5000 
octant_analysis(mod)

end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))