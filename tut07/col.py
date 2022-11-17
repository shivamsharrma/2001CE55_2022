import openpyxl
wb = openpyxl.load_workbook('output.xlsx')
sheet = wb.active
from openpyxl.styles import PatternFill,Border,Side
for rows in sheet.iter_rows(min_row=2, max_row=7, min_col=22, max_col=29):
    yellow = "00FFFF00"
    for cell in rows:
        if cell.value==1:
            cell.fill = PatternFill(start_color=yellow, end_color=yellow,fill_type = "solid")
a=Side(border_style='thin',color="00000000")
rownum=6
colnum=19
rowloc=1
colloc=12
border=Border(top=a,bottom=a,left=a,right=a)
for i in range (rowloc, rowloc+rownum):
    for j in range (colloc,colnum+colloc) :
        sheet.cell(row=i+1, column=j+1).border=border
        if i==rowloc:
            sheet. cell (row=i+1,column=j+1).border=border
        if i==rowloc+rownum-1:
            sheet.cell(row=i+1,column=j+1).border=border
rownum=9
colnum=3
rowloc=8
colloc=27
border=Border(top=a,bottom=a,left=a,right=a)
for i in range (rowloc, rowloc+rownum):
    for j in range (colloc,colnum+colloc) :
        sheet.cell(row=i+1, column=j+1).border=border
        if i==rowloc:
            sheet. cell (row=i+1,column=j+1).border=border
        if i==rowloc+rownum-1:
            sheet.cell(row=i+1,column=j+1).border=border
rownum=9
colnum=3
rowloc=0
colloc=43
border=Border(top=a,bottom=a,left=a,right=a)
for i in range (rowloc, rowloc+rownum):
    for j in range (colloc,colnum+colloc) :
        sheet.cell(row=i+1, column=j+1).border=border
        if i==rowloc:
            sheet. cell (row=i+1,column=j+1).border=border
        if i==rowloc+rownum-1:
            sheet.cell(row=i+1,column=j+1).border=border
 # rownum=25+vary
rownum=25
colnum=3
rowloc=0
colloc=47
border=Border(top=a,bottom=a,left=a,right=a)
for i in range (rowloc, rowloc+rownum):
    for j in range (colloc,colnum+colloc) :
        sheet.cell(row=i+1, column=j+1).border=border
        if i==rowloc:
            sheet. cell (row=i+1,column=j+1).border=border
        if i==rowloc+rownum-1:
            sheet.cell(row=i+1,column=j+1).border=border
rownum=9
colnum=9
rowloc=0
colloc=33
dis=5
#vary
T=4+1
for k in range(T):
    for i in range (rowloc, rowloc+rownum):
        for j in range (colloc,colnum+colloc) :
            sheet.cell(row=i+1, column=j+1).border=border
            if i==rowloc:
                sheet. cell (row=i+1,column=j+1).border=border
            if i==rowloc+rownum-1:
                sheet.cell(row=i+1,column=j+1).border=border
    rowloc+=rownum+dis
rownum=8
colnum=8
rowloc=2
colloc=35
dis=6
#vary
T=4+1
l=[[],[],[],[],[],[],[],[]]
for k in range(T):
    for i in range (rowloc, rowloc+rownum):
        l=[[],[],[],[],[],[],[],[]]
        x=0
        for j in range (colloc,colnum+colloc) :
            l[x].append(sheet.cell(row=i,column=j).value)
            l[x].append(i)
            l[x].append(j)
            x+=1
            a=max(l)
            print(l)
        sheet.cell(row=a[1],column=a[2]).fill=PatternFill(start_color=yellow, end_color=yellow,fill_type = "solid")
    rowloc+=rownum+dis
wb.save("color.xlsx")